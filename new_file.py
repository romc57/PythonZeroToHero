import random
from itertools import cycle
from openpyxl import Workbook
from collections import defaultdict

# Setup
professions = ["nurse", "doctor", "therapist", "psychologist", "care_assistant"]
days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Sunday"]
hours = list(range(8, 18))
names_pool = ["Alice", "Bob", "Charlie", "Diana", "Eli", "Fiona", "George", "Hannah", "Ivan", "Julia"]

# Step 1: Generate caretakers with rotating professions
caretakers = []
used_names = set()
profession_cycle = cycle(professions)

for _ in range(8):
    while True:
        name = random.choice(names_pool)
        if name not in used_names:
            used_names.add(name)
            break
    profession = next(profession_cycle)
    working_days = random.sample(days, random.randint(3, 6))
    block_length = random.randint(4, 8)
    start_hour = random.randint(8, 17 - block_length + 1)
    working_hours = list(range(start_hour, start_hour + block_length))
    caretakers.append({
        "name": f"{name} ({profession})",
        "profession": profession,
        "working_days": working_days,
        "working_hours": working_hours
    })

# Step 2: Assign patients to caretakers
wb = Workbook()
wb.remove(wb.active)
patient_counter = 1
patient_assignments = {}

def get_patient_id():
    global patient_counter
    pid = f"P{patient_counter:03d}"
    patient_counter += 1
    return pid

for ct in caretakers:
    ws = wb.create_sheet(title=ct["name"][:31])
    ws.cell(row=1, column=1, value="Hour")
    for col, day in enumerate(days, 2):
        ws.cell(row=1, column=col, value=day)
    for row, hour in enumerate(hours, 2):
        ws.cell(row=row, column=1, value=hour)

    available_slots = [(d, h) for d in ct["working_days"] for h in ct["working_hours"]]
    used_slots = set()

    while available_slots:
        pid = get_patient_id()
        assigned_days = set()
        slots_to_assign = []

        for d, h in available_slots:
            if d not in assigned_days:
                slots_to_assign.append((d, h))
                assigned_days.add(d)
            if len(slots_to_assign) >= random.randint(3, 5):
                break

        for d, h in slots_to_assign:
            row = hours.index(h) + 2
            col = days.index(d) + 2
            ws.cell(row=row, column=col, value=pid)
            patient_assignments.setdefault(pid, []).append((d, h, ct["profession"]))
            used_slots.add((d, h))

        available_slots = [slot for slot in available_slots if slot not in used_slots]

# Step 3: Save caretaker schedule
wb.save("fixed_caretaker_schedule_unique_days.xlsx")

# Step 4: Create patient-based schedule (hour x day)
patient_wb = Workbook()
patient_wb.remove(patient_wb.active)
patient_grid_map = defaultdict(lambda: [["" for _ in days] for _ in hours])

for pid, entries in patient_assignments.items():
    for d, h, prof in entries:
        r = hours.index(h)
        c = days.index(d)
        patient_grid_map[pid][r][c] = prof

for pid, grid in patient_grid_map.items():
    ws = patient_wb.create_sheet(title=pid)
    ws.cell(row=1, column=1, value="Hour")
    for col, day in enumerate(days, 2):
        ws.cell(row=1, column=col, value=day)
    for row, hour in enumerate(hours, 2):
        ws.cell(row=row, column=1, value=hour)
        for col, day in enumerate(days, 2):
            ws.cell(row=row, column=col, value=grid[row - 2][col - 2])

# Step 5: Save patient schedule
patient_wb.save("fixed_patient_schedule_unique_days.xlsx")
